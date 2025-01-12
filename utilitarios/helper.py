import os 
import re
import openpyxl
import pandas as pd
import configparser
from utilitarios.logger_config import logger

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def ler_configuracao(caminho_config:str):
    """Lê o arquivo de configuração."""

    config = configparser.ConfigParser()
    config.read(caminho_config)
    return config

def listar_subpastas(caminho_pasta:str):
    """Lista as subpastas dentro do caminho especificado."""

    return [
        nome for nome in os.listdir(caminho_pasta) 
        if os.path.isdir(os.path.join(caminho_pasta, nome))
    ]

def ler_arquivo_excel(caminho_excel:str):
    """Lê o arquivo Excel especificado."""
    try:
        df = pd.read_excel(caminho_excel)
        return df
    except Exception as e:
        logger.error(f"Erro ao ler o arquivo Excel: {e}")
        return None

def salvar_dados(dados: list[dict], empresa: str, tipo_arquivo: str):
    """
    Salva os dados no formato desejado.

    Args:
        df (pd.DataFrame): DataFrame com os dados a serem salvos.
        empresa (str): Nome da empresa.
        tipo_arquivo (str): Formato de arquivo desejado (csv, xlsx).
    """

    df = pd.DataFrame(dados)

    if tipo_arquivo == "csv":
        df.to_csv(f"{empresa}_dados.csv", index=False)
    elif tipo_arquivo == "xlsx":
        df.to_excel(f"{empresa}_dados.xlsx", index=False)
    elif tipo_arquivo == "json":
        df.to_json(f"{empresa}_dados.json", orient="records")
    elif tipo_arquivo == "xml":
        df.to_xml(f"{empresa}_dados.xml", root_name="dados")
    elif tipo_arquivo == "html":
        df.to_html(f"{empresa}_dados.html", index=False)

    else:
        logger.error(f"Formato de arquivo inválido: {tipo_arquivo}")

def find_previous_gto(procedure_start, gto_matches):
    for gto in reversed(gto_matches):
        if gto.end() < procedure_start:
            return gto.groups()
    return (None, None)

def obter_conteudo_parenteses(texto):
    ultimo_parentese = re.search(r'\(([^)]*)\)', texto)
    conteudo_parentese = ultimo_parentese.group(1) if ultimo_parentese else None
    return conteudo_parentese

def criar_planilha_inicial(arquivo: str):
    """
    Cria a planilha inicial com o cabeçalho formatado e bordas.

    Args:
        arquivo (str): Caminho e nome do arquivo de planilha.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plano"

    # Configuração da borda padrão
    borda = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    # Configuração do cabeçalho principal (linha 1)
    sheet.merge_cells("A1:D1")
    sheet.merge_cells("E1:K1")

    celula_nome_dentista = sheet["A1"]
    celula_nome_dentista.value = "NOME DO DENTISTA: COLOCA SEU NOME"
    celula_nome_dentista.font = Font(bold=True, color="000000")
    celula_nome_dentista.alignment = Alignment(horizontal="center", vertical="center")
    celula_nome_dentista.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    celula_mes_referencia = sheet["E1"]
    celula_mes_referencia.value = "MÊS DE REFERENCIA: JANEIRO ( COLOCA MÊS DE REFERENCIA )"
    celula_mes_referencia.font = Font(bold=True, color="000000")
    celula_mes_referencia.alignment = Alignment(horizontal="center", vertical="center")
    celula_mes_referencia.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in sheet["A1:D1"]:
        for cell in row:
            cell.font = Font(bold=True, color="FF0000")
            cell.border = borda

    for row in sheet["E1:K1"]:
        for cell in row:
            cell.font = Font(bold=True, color="FF0000")
            cell.border = borda

    sheet.row_dimensions[1].height = 15
    sheet.row_dimensions[1].height *= 3

    # Configuração do cabeçalho (linha 2)
    cabecalho = [
        "DATA",
        "NOME DO PACIENTE",
        "PLANO",
        "PROCEDIMENTO REALIZADO",
        "DENTE",
        "FACE",
        "REGIÃO",
        "NUMERO DA GTO",
        "PAGO / GLOSADO",
        "VALOR PAGO",
        "REPASSE DENTISTA",
    ]

    for col_num, texto in enumerate(cabecalho, start=1):
        celula = sheet.cell(row=2, column=col_num)
        celula.value = texto
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    # Configuração da célula de total (coluna M)
    sheet["M2"].value = "TOTAL"
    sheet["M2"].font = Font(bold=True)
    sheet["M2"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["M2"].fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid")
    sheet["M2"].border = borda

    sheet.merge_cells("M3:M4")
    total_cell = sheet["M3"]
    total_cell.value = "=SUM(K3:K1002)"
    total_cell.number_format = "R$ #,##0.00"
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    total_cell.border = borda

    sheet.row_dimensions[2].height = 15
    sheet.row_dimensions[2].height *= 2

    # Salvar o arquivo
    workbook.save(arquivo)

def salvar_dados_planilha(dados: list[dict], plano: str, arquivo: str):
    """
    Salva os dados na planilha especificada, aplicando bordas nas novas linhas.

    Args:
        dados (list): Dados a serem salvos.
        plano (str): Nome do plano.
        arquivo (str): Caminho e nome do arquivo de planilha.
    """

    try:
        # Carregar o arquivo de planilha
        workbook = openpyxl.load_workbook(arquivo)
        sheet = workbook.active

        borda = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        colunas = {
            'Data de Realização': 'DATA',
            'Nome do Beneficiário': 'NOME DO PACIENTE',
            'Plano': 'PLANO',
            'Nome Procedimento': 'PROCEDIMENTO REALIZADO',
            'Dente/Região': 'DENTE',
            'Face': 'FACE',
            'Regiao': 'REGIÃO',
            'GTO': 'NUMERO DA GTO',
            'Valor Glosa': 'PAGO / GLOSADO',
            'Valor Processado': 'VALOR PAGO',
        }

        # Identificar a próxima linha vazia para inserir os dados
        ultima_linha = sheet.max_row
        proxima_linha = ultima_linha + 1

        # Adicionar os dados na planilha
        for dado in dados:
            dado['Plano'] = plano
            if dado['Valor Processado']:
                dado['Valor Processado'] = dado['Valor Processado'].replace('.', ',')
            if dado['Valor Glosa']:
               dado['Valor Glosa'] = dado['Valor Glosa'].replace('.', ',')
            nova_linha = [dado.get(key, None) for key in colunas.keys()]
            for col_index, valor in enumerate(nova_linha, start=1):
                cell = sheet.cell(row=proxima_linha, column=col_index, value=valor)
                cell.alignment = Alignment(horizontal="center")
                cell.border = borda

            # Fórmulas e estilos específicos
            sheet.cell(row=proxima_linha, column=11).value = f"=J{proxima_linha}*35/100"
            sheet.cell(row=proxima_linha, column=11).border = borda
            sheet.cell(row=proxima_linha, column=9).fill = PatternFill(start_color="FF8669", end_color="FF8669", fill_type="solid")
            sheet.cell(row=proxima_linha, column=10).fill = PatternFill(start_color="C2F3B7", end_color="C2F3B7", fill_type="solid")
            sheet.cell(row=proxima_linha, column=11).fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

            proxima_linha += 1

        # Salvar o arquivo atualizado
        workbook.save(arquivo)
        logger.info(f"Relatório salvo com sucesso em {arquivo}")

    except Exception as e:
        logger.error(f"Erro ao salvar relatório: {e}")

